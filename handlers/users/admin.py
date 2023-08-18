import asyncio
from aiogram import types, utils
from aiogram.dispatcher import FSMContext
from aiogram.types import Message, ReplyKeyboardRemove

from data.config import ADMINS
from loader import dp, db, bot
from keyboards.default.admin import admin_menu
from keyboards.default.menu import main

import sqlite3
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime, timedelta

connection = sqlite3.connect('data/main.db')
cursor = connection.cursor()

# admin uchun excel faylga yozib beradigan funksiya yozamiz
@dp.message_handler(commands='admin', user_id=ADMINS)
async def admins(message:Message):
    await message.answer('admin menu ochildi qaytish uchun /start buyruguni bering',reply_markup=admin_menu) #endi buttonlarni tayyorlab olamiz admin uchun


@dp.message_handler(text='📑 Barcha ma\'lumotlari olish')
async def all_date(message: Message):
    #malumotlarni hammasini olish
    cursor.execute("SELECT * FROM register")
    rows = cursor.fetchall()

    #excel faylga endi yozamiz malumotlarni
    wb = Workbook()
    ws = wb.active
    ws.title ='Umumiy malumotlar'

    ws.cell(row=1, column=1).value='ID'
    ws.cell(row=1, column=2).value='Ism Fam'
    ws.cell(row=1, column=3).value='Tugilgan sana'
    ws.cell(row=1, column=4).value='Tel raqami'
    ws.cell(row=1, column=5).value='Manzil'
    ws.cell(row=1, column=6).value='Yunalish'
    ws.cell(row=1, column=7).value='Bosh kunlari'
    ws.cell(row=1, column=8).value='Bosh vaqtlari'
    ws.cell(row=1, column=9).value='Qoshimcha malumot'
    ws.cell(row=1, column=10).value='Royxatdan otgan sana'


    for row in rows:
        ws.append(row)

    #file ni yuborish
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    document = types.InputFile(file, filename='umumiy.xlsx')
    await message.answer_document(document=document)
    wb.close()



@dp.message_handler(text='📎 Oylik')
async def monthly(message: Message):
    wb = Workbook()
    ws = wb.active
    sana = datetime.now().date() - timedelta(days=datetime.now().day - 1)
    query = f"SELECT * FROM register WHERE created_at >= '{sana}'"
    cursor.execute(query)
    data = cursor.fetchall()

    #faylga malumotni yozish
    ws.title = 'Oylik malumot'

    # ustun nomlari
    ws.cell(row=1, column=1).value = 'ID'
    ws.cell(row=1, column=2).value = 'Ism Fam'
    ws.cell(row=1, column=3).value = 'Tugilgan sana'
    ws.cell(row=1, column=4).value = 'Tel raqami'
    ws.cell(row=1, column=5).value = 'Manzil'
    ws.cell(row=1, column=6).value = 'Yunalish'
    ws.cell(row=1, column=7).value = 'Bosh kunlari'
    ws.cell(row=1, column=8).value = 'Bosh vaqtlari'
    ws.cell(row=1, column=9).value = 'Qoshimcha malumot'
    ws.cell(row=1, column=10).value = 'Royxatdan otgan sana'

    for row in data:
        ws.append(row)

    filename = f'malumotlar_{sana:%Y-%m}.xlsx' #bu bizga fayl yuborilganda fayl nomida yil va oy berilgan boladi
    wb.save(filename)
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    document = types.InputFile(file, filename=filename)
    await message.answer_document(document=document)
    wb.close()

    # boldi oylik malumotni ham bitirib oldik

@dp.message_handler(text='📎 Haftalik')
async def weekly(message: Message):
    wb = Workbook()
    ws = wb.active
    query = f"SELECT * FROM register WHERE created_at >= date('now', '-7 day')"
    cursor.execute(query)
    data = cursor.fetchall()

    #faylga malumotni yozish
    ws.title = f'Haftalik malumotlar'

    # ustun nomlari
    ws.cell(row=1, column=1).value = 'ID'
    ws.cell(row=1, column=2).value = 'Ism Fam'
    ws.cell(row=1, column=3).value = 'Tugilgan sana'
    ws.cell(row=1, column=4).value = 'Tel raqami'
    ws.cell(row=1, column=5).value = 'Manzil'
    ws.cell(row=1, column=6).value = 'Yunalish'
    ws.cell(row=1, column=7).value = 'Bosh kunlari'
    ws.cell(row=1, column=8).value = 'Bosh vaqtlari'
    ws.cell(row=1, column=9).value = 'Qoshimcha malumot'
    ws.cell(row=1, column=10).value = 'Royxatdan otgan sana'

    for row in data:
        ws.append(row)

    filename = f'haftalik.xlsx'
    wb.save(filename)
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    document = types.InputFile(file, filename=filename)
    await message.answer_document(document=document)
    wb.close()



@dp.message_handler(text='📎 Kunlik')
async def day(message: Message):
    wb = Workbook()
    ws = wb.active
    today = datetime.today().strftime('%Y-%m-%d')
    query = f"SELECT * FROM register WHERE created_at >= date('now', '-0 day')"
    cursor.execute(query)
    data = cursor.fetchall()

    #faylga malumotni yozish
    ws.title = f'Kunlik malumotlar {today}'

    # ustun nomlari
    ws.cell(row=1, column=1).value = 'ID'
    ws.cell(row=1, column=2).value = 'Ism Fam'
    ws.cell(row=1, column=3).value = 'Tugilgan sana'
    ws.cell(row=1, column=4).value = 'Tel raqami'
    ws.cell(row=1, column=5).value = 'Manzil'
    ws.cell(row=1, column=6).value = 'Yunalish'
    ws.cell(row=1, column=7).value = 'Bosh kunlari'
    ws.cell(row=1, column=8).value = 'Bosh vaqtlari'
    ws.cell(row=1, column=9).value = 'Qoshimcha malumot'
    ws.cell(row=1, column=10).value = 'Royxatdan otgan sana'

    for row in data:
        ws.append(row)

    filename = f'Kun {today}.xlsx' #bu bizga fayl yuborilganda fayl nomida yil va oy berilgan boladi
    wb.save(filename)
    file = BytesIO()
    wb.save(file)
    file.seek(0)
    document = types.InputFile(file, filename=filename)
    await message.answer_document(document=document)
    wb.close()



@dp.message_handler(text='🔄 Asosiy sahifa')
async def main_(message: Message):
    await message.answer('Bosh menu', reply_markup=main)





@dp.message_handler(text="🎯 Reklama", user_id=ADMINS, state=None)
async def send_ad_to_all(message: types.Message, state=FSMContext):
    try:
        await message.answer(
            "✅ Reklamani botga forward qiling. Siz forward qilgan reklama to'gridan to'g'ri barcha foydalanuchilarga yuboriladi. \n\n✅ Yoki xabar matnini kiriting: \n\n⚠️ Xabar yuborishni istamasangiz /bekor kamandasini kiriting.",
            reply_markup=ReplyKeyboardRemove())
        # print("1")
    except Exception as e:
        await message.answer(
            "✅ Reklamani botga forward qiling. Siz forward qilgan reklama to'gridan to'g'ri barcha foydalanuchilarga yuboriladi. \n\n✅ Yoki xabar matnini kiriting: \n\n⚠️ Xabar yuborishni istamasangiz /bekor kamandasini kiriting.")
        # print("2")

    await state.set_state("send_users")


@dp.message_handler(state="send_users", text="/bekor")
async def send_ad_to_all(message: types.Message, state=FSMContext):
    await message.answer("❌ Bekor qilindi")
    await state.finish()


@dp.message_handler(state="send_users", content_types="text", is_forwarded=False)
async def send_ad_to_all(message: types.Message, state=FSMContext):
    await state.finish()
    bloklaganlar = 0
    jonlilar = 0

    users = db.select_all_users()
    for user in users:
        user_id = user[0]
        try:
            await bot.send_message(user_id, message.text)
            jonlilar += 1

        except utils.exceptions.BotBlocked as e:
            bloklaganlar += 1

        except:
            pass

        else:
            pass
        await asyncio.sleep(0.05)
    await message.answer("✅ Xabaringiz foydalanuvchilarga yetkazildi...")
    await message.answer(f"Jami foydalanuvchilar soni: {bloklaganlar + jonlilar}")
    await message.answer(f"Bloklagan foydalanuvchilar soni: {bloklaganlar}")
    await message.answer(f"Faol foydalanuvchilar soni: {jonlilar}")


@dp.message_handler(state="send_users", content_types="any", is_forwarded=True)
async def send_ad_to_all(message: types.Message, state=FSMContext):
    await state.finish()

    bloklaganlar = 0
    jonlilar = 0

    users = db.select_all_users()
    for user in users:
        user_id = user[0]

        try:
            await message.forward(user_id, message.forward_from.id, message.forward_from_message_id)
            jonlilar += 1
        except utils.exceptions.BotBlocked as e:
            bloklaganlar += 1

        except:
            await message.forward(user_id, message.forward_from_chat.id, message.forward_from_message_id)
            jonlilar += 1
        else:
            pass

        await asyncio.sleep(0.05)

    await message.answer("✅ Xabaringiz foydalanuvchilarga yetkazildi...")

    await message.answer(f"Jami foydalanuvchilar soni: {bloklaganlar + jonlilar}")
    await message.answer(f"Bloklagan foydalanuvchilar soni: {bloklaganlar}")
    await message.answer(f"Faol foydalanuvchilar soni: {jonlilar}")




@dp.message_handler(text="/allusers", user_id=ADMINS)
async def get_all_users(message: types.Message):
    users = db.select_all_users()
    print(users[0][0])
    await message.answer(users)



@dp.message_handler(text="/cleandb", user_id=ADMINS)
async def get_all_users(message: types.Message):
    db.delete_users()
    await message.answer("Baza tozalandi!")




